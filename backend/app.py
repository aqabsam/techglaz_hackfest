from flask import Flask, Response, jsonify, request, send_file
from werkzeug.utils import secure_filename
from datetime import datetime

try:
    from flask_cors import CORS
except ImportError:  # pragma: no cover - fallback for environments without the optional package
    class CORS:  # type: ignore[override]
        def __init__(self, *args, **kwargs):
            pass

        def init_app(self, *args, **kwargs):
            return None
import json
import os
import signal
import subprocess
import sys
import threading
import time
import urllib.error
import urllib.request
import uuid
import re

app = Flask(__name__)

try:
    CORS(app)
except Exception:  # pragma: no cover - fallback for environments without the optional package
    pass


@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, PATCH, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    return response


@app.route('/api/teacher-credentials', methods=['OPTIONS'])
def teacher_credentials_options():
    return jsonify({}), 200


@app.route('/api/students', methods=['OPTIONS'])
def students_options():
    return jsonify({}), 200


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STUDENT_PHOTO_DIR = os.path.join(BASE_DIR, 'KnownFaces')
STUDENTS_FILE = os.path.join(BASE_DIR, 'students.json')
TEACHER_PROFILE_FILE = os.path.join(BASE_DIR, 'teacherCredentials.json')
EXPORT_DIR = os.environ.get('ATTENDANCE_EXPORT_DIR', os.path.join(BASE_DIR, 'exports'))
os.makedirs(EXPORT_DIR, exist_ok=True)
EXCEL_FILE = os.environ.get('ATTENDANCE_EXPORT_FILE') or os.path.join(EXPORT_DIR, 'Attendance.xlsx')
ATTENDANCE_SCRIPT = os.path.join(BASE_DIR, 'attendance_runner.py')
ATTENDANCE_RUNTIME_DIR = os.path.join(BASE_DIR, 'runtime')
ATTENDANCE_FRAME_FILE = os.path.join(ATTENDANCE_RUNTIME_DIR, 'attendance_frame.jpg')
HOST = os.environ.get('HOST', '0.0.0.0')
PORT = int(os.environ.get('PORT', '5000'))
USE_FIREBASE = os.environ.get('USE_FIREBASE', '').strip().lower() in {'1', 'true', 'yes', 'on'}

os.makedirs(STUDENT_PHOTO_DIR, exist_ok=True)
os.makedirs(ATTENDANCE_RUNTIME_DIR, exist_ok=True)

attendance_process = None
attendance_last_error = ''


def load_env_file(path):
    if not os.path.exists(path):
        return

    try:
        with open(path, 'r', encoding='utf-8') as file:
            for raw_line in file:
                line = raw_line.strip()
                if not line or line.startswith('#') or '=' not in line:
                    continue

                key, value = line.split('=', 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")

                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception:
        pass


load_env_file(os.path.join(os.path.dirname(BASE_DIR), '.env'))
load_env_file(os.path.join(os.path.dirname(BASE_DIR), '.env.local'))


FIREBASE_DATABASE_URL = os.environ.get('FIREBASE_DATABASE_URL') or os.environ.get('VITE_FIREBASE_DATABASE_URL') or 'https://attenzo-web-default-rtdb.firebaseio.com'
current_excel_file = EXCEL_FILE


def clear_latest_frame():
    try:
        if os.path.exists(ATTENDANCE_FRAME_FILE):
            os.remove(ATTENDANCE_FRAME_FILE)
    except Exception:
        pass


def stream_latest_frame():
    boundary = b'--frame'
    last_frame_mtime = 0.0

    while attendance_process and attendance_process.poll() is None:
        if not os.path.exists(ATTENDANCE_FRAME_FILE):
            time.sleep(0.1)
            continue

        try:
            mtime = os.path.getmtime(ATTENDANCE_FRAME_FILE)
            if mtime <= last_frame_mtime:
                time.sleep(0.08)
                continue

            last_frame_mtime = mtime
            with open(ATTENDANCE_FRAME_FILE, 'rb') as file:
                frame = file.read()

            if not frame:
                time.sleep(0.08)
                continue

            yield (
                boundary + b'\r\n'
                b'Content-Type: image/jpeg\r\n'
                b'Cache-Control: no-cache, no-store, must-revalidate\r\n\r\n' + frame + b'\r\n'
            )
            time.sleep(0.08)
        except Exception:
            time.sleep(0.1)

    yield b''


hidden_name_fragments = ['mohd aqab sami', 'mohd aqabsami', 'aban shami']
hidden_phone_digits = {'8986392298'}


def normalize_student_value(value):
    return str(value or '').strip().lower().replace('\n', ' ').replace('\t', ' ')


def digits_only(value):
    return ''.join(ch for ch in str(value or '') if ch.isdigit())


def is_visible_student(student):
    normalized_name = normalize_student_value(student.get('name', ''))
    normalized_phone = digits_only(student.get('phoneNumber', ''))
    return not any(fragment in normalized_name for fragment in hidden_name_fragments) and normalized_phone not in hidden_phone_digits


def filter_visible_students(students):
    return [student for student in students if is_visible_student(student)]


def firebase_url(path):
    cleaned_path = path.strip('/')
    base_url = FIREBASE_DATABASE_URL.rstrip('/')
    return f'{base_url}/{cleaned_path}.json' if cleaned_path else f'{base_url}.json'


def firebase_request(path, method='GET', payload=None):
    data = None if payload is None else json.dumps(payload).encode('utf-8')
    headers = {'Content-Type': 'application/json'} if payload is not None else {}
    request_obj = urllib.request.Request(firebase_url(path), data=data, method=method, headers=headers)
    with urllib.request.urlopen(request_obj, timeout=10) as response:
        body = response.read().decode('utf-8').strip()
        return json.loads(body) if body else None


def load_local_students():
    if os.path.exists(STUDENTS_FILE):
      try:
        with open(STUDENTS_FILE, 'r', encoding='utf-8') as file:
          students = json.load(file)
          students = students if isinstance(students, list) else []
          return filter_visible_students(students)
      except Exception:
        return []
    return []


def write_local_students(students):
    with open(STUDENTS_FILE, 'w', encoding='utf-8') as file:
        json.dump(students, file, indent=2, ensure_ascii=False)


def load_teacher_profile():
    if os.path.exists(TEACHER_PROFILE_FILE):
        try:
            with open(TEACHER_PROFILE_FILE, 'r', encoding='utf-8') as file:
                data = json.load(file)
                return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    return {}


def save_teacher_profile(profile):
    with open(TEACHER_PROFILE_FILE, 'w', encoding='utf-8') as file:
        json.dump(profile if isinstance(profile, dict) else {}, file, indent=2, ensure_ascii=False)


def student_photo_source_exists(student):
    photo_url = str(student.get('photoUrl', '') or '').strip()
    if photo_url:
        return True

    photo_filename = str(student.get('photoFilename', '') or '').strip()
    if photo_filename:
        return os.path.exists(os.path.join(STUDENT_PHOTO_DIR, photo_filename))

    return False


def collect_photo_ready_students(students):
    ready_students = [student for student in students if student_photo_source_exists(student)]
    missing_students = [student for student in students if not student_photo_source_exists(student)]
    return ready_students, missing_students


def monitor_attendance_process(process):
    global attendance_process, attendance_last_error

    output_lines = []

    try:
        if process.stdout:
            for raw_line in iter(process.stdout.readline, ''):
                if not raw_line:
                    break
                line = raw_line.rstrip()
                if line:
                    output_lines.append(line)
                    output_lines[:] = output_lines[-8:]
    except Exception:
        pass

    return_code = process.wait()

    if return_code == 0:
        attendance_last_error = ''
    else:
        attendance_last_error = '\n'.join(output_lines[-4:]) if output_lines else 'Attendance process exited before it could stay running.'

    if attendance_process is process:
        attendance_process = None


def normalize_student_records(records):
    if not records:
        return []
    if isinstance(records, list):
        return [student for student in records if isinstance(student, dict)]
    if isinstance(records, dict):
        return [student for student in records.values() if isinstance(student, dict)]
    return []


def load_students_from_firebase():
    if not USE_FIREBASE:
        return []

    try:
        records = firebase_request('students')
        students = filter_visible_students(normalize_student_records(records))
        if students:
            try:
                write_local_students(students)
            except Exception:
                pass
            return students
    except Exception:
        pass

    return load_local_students()


def save_students(students):
    visible_students = filter_visible_students(students)
    write_local_students(visible_students)

    payload = {student.get('id'): student for student in visible_students if student.get('id')}
    try:
        firebase_request('students', method='PUT', payload=payload)
    except Exception:
        pass


def load_students():
    try:
        local_students = load_local_students()
        if local_students:
            return local_students
    except Exception:
        pass

    try:
        students = load_students_from_firebase()
        if students:
            return students
    except Exception:
        pass

    return load_local_students()


def student_photo_url(photo_filename):
    return f'{request.host_url.rstrip("/")}/api/student-photo/{photo_filename}'


def serialize_student(student):
    photo_filename = student.get('photoFilename', '')
    photo_url = student.get('photoUrl', '')
    return {
        'id': student.get('id', ''),
        'name': student.get('name', ''),
        'rollNumber': student.get('rollNumber', ''),
        'phoneNumber': student.get('phoneNumber', ''),
        'photoUrl': photo_url or (student_photo_url(photo_filename) if photo_filename else ''),
        'photoFilename': photo_filename,
        'createdAt': student.get('createdAt', ''),
    }


def create_export_filename(students):
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    roll_numbers = [str(student.get('rollNumber', '')).strip() for student in students if student.get('rollNumber')]
    unique_rolls = []
    for roll_number in roll_numbers:
        if roll_number not in unique_rolls:
            unique_rolls.append(roll_number)

    if unique_rolls:
        roll_label = '_'.join(re.sub(r'[^A-Za-z0-9]+', '', roll_number) or 'roll' for roll_number in unique_rolls[:8])
        if len(unique_rolls) > 8:
            roll_label = f"{roll_label}_{len(unique_rolls)}students"
    else:
        roll_label = 'AllStudents'

    return os.path.join(EXPORT_DIR, f"Attenzo_{roll_label}_{timestamp}.xlsx")


def get_active_excel_file():
    return current_excel_file if current_excel_file else EXCEL_FILE


def create_session_excel_file():
    return get_active_excel_file() or os.path.join(EXPORT_DIR, 'Attendance.xlsx')


def normalize_source(mode, source):
    if mode == 'webcam':
        return '0'
    if not source:
        return ''
    return source.strip()


def normalize_status_value(value):
    return str(value or '').strip().lower()


def is_status_value(value):
    return normalize_status_value(value) in {'present', 'absent'}


def is_non_counting_status(value):
    return normalize_status_value(value) in {'holiday', 'sunday'}


def get_class_start_date(profile=None):
    profile = profile or load_teacher_profile()
    return str(profile.get('classStartDate', '') or '').strip()


def get_day_overrides(profile=None):
    profile = profile or load_teacher_profile()
    overrides = profile.get('dayOverrides', {}) or {}
    if isinstance(overrides, dict):
        return {str(date_key): str(status_value) for date_key, status_value in overrides.items() if str(date_key).strip()}
    return {}


def filter_records_by_class_start(records, profile=None):
    start_date = get_class_start_date(profile)
    if not start_date:
        return records
    return [record for record in records if not record.get('date') or record.get('date') >= start_date]


def parse_attendance_row(row):
    if not row:
        return None

    values = list(row)
    if len(values) >= 6:
        return {
            'name': values[0] or '',
            'rollNumber': values[1] or '',
            'status': values[2] or '',
            'checkinTime': str(values[3]) if values[3] else '',
            'date': str(values[4]) if values[4] else '',
            'day': values[5] or '',
        }

    if len(values) >= 5:
        return {
            'name': values[0] or '',
            'rollNumber': '',
            'checkinTime': str(values[1]) if values[1] else '',
            'status': values[2] or '',
            'date': str(values[3]) if values[3] else '',
            'day': values[4] or '',
        }

    if len(values) >= 4:
        if is_status_value(values[3]):
            return {
                'name': values[1] or '',
                'rollNumber': values[0] or '',
                'status': values[3] or '',
                'checkinTime': str(values[2]) if values[2] else '',
                'date': '',
                'day': '',
            }

        return {
            'name': values[0] or '',
            'rollNumber': values[1] or '',
            'status': values[2] or '',
            'checkinTime': str(values[3]) if values[3] else '',
            'date': '',
            'day': '',
        }

    return None


def parse_sheet_day(sheet_title, title_value=''):
    title_text = str(title_value or '')
    match = re.search(r'\(([^)]+)\)', title_text)
    if match:
        return match.group(1)

    return ''


def get_today_sheet_name():
    return datetime.now().strftime('%Y-%m-%d')


def today_sheet_has_records(excel_file):
    if not os.path.exists(excel_file):
        return False

    try:
        from openpyxl import load_workbook

        wb = load_workbook(excel_file)
        sheet_name = get_today_sheet_name()
        if sheet_name not in wb.sheetnames:
            return False

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row and any(cell not in (None, '') for cell in row):
                return True
    except Exception:
        return False

    return False


def read_sheet_records(ws):
    title_value = ws['A1'].value or ''
    day_hint = parse_sheet_day(ws.title, title_value)
    date_hint = ws.title
    records = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        record = parse_attendance_row(row)
        if not record:
            continue

        if not record.get('date'):
            record['date'] = date_hint
        if not record.get('day'):
            record['day'] = day_hint
        records.append(record)

    return records


def persist_day_override(excel_file, day_date, status, students):
    from openpyxl import Workbook, load_workbook

    try:
        datetime.strptime(day_date, '%Y-%m-%d')
    except ValueError:
        return False

    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)

    if day_date in workbook.sheetnames:
        workbook.remove(workbook[day_date])

    worksheet = workbook.create_sheet(title=day_date)
    day_name = datetime.strptime(day_date, '%Y-%m-%d').strftime('%A')
    worksheet['A1'] = f'Attendance - {day_date} ({day_name})'
    worksheet['A2'] = 'Name'
    worksheet['B2'] = 'Roll Number'
    worksheet['C2'] = 'Status'
    worksheet['D2'] = 'Check-in Time'
    worksheet['E2'] = 'Date'
    worksheet['F2'] = 'Day'

    for index, student in enumerate(students, start=3):
        worksheet[f'A{index}'] = student.get('name', '')
        worksheet[f'B{index}'] = student.get('rollNumber', '')
        worksheet[f'C{index}'] = status
        worksheet[f'D{index}'] = ''
        worksheet[f'E{index}'] = day_date
        worksheet[f'F{index}'] = day_name

    workbook.save(excel_file)
    return True


def list_attendance_files():
    files = []
    if not os.path.isdir(EXPORT_DIR):
        return files

    for filename in sorted(os.listdir(EXPORT_DIR)):
        full_path = os.path.join(EXPORT_DIR, filename)
        if os.path.isfile(full_path) and filename.lower().endswith('.xlsx'):
            files.append(full_path)

    return files


def collect_attendance_records():
    records = []
    seen = set()

    for excel_file in list_attendance_files():
        if not os.path.exists(excel_file):
            continue

        try:
            from openpyxl import load_workbook

            wb = load_workbook(excel_file, data_only=True)
            for ws in wb.worksheets:
                for record in read_sheet_records(ws):
                    key = (
                        record.get('date', ''),
                        record.get('name', ''),
                        record.get('rollNumber', ''),
                        record.get('checkinTime', ''),
                        record.get('status', ''),
                    )
                    if key not in seen:
                        seen.add(key)
                        records.append(record)
        except Exception:
            continue

    records = filter_records_by_class_start(records)
    records.sort(key=lambda item: (item.get('date', ''), item.get('day', ''), item.get('name', '')))
    return records


@app.route('/')
def health_check():
    return jsonify({
        'status': 'ok',
        'service': 'attenzo-backend',
        'routes': ['/api/students', '/api/attendance/start', '/api/reports']
    })


@app.route('/api/students', methods=['GET'])
def get_students():
    return jsonify([serialize_student(student) for student in load_students()])


@app.route('/api/students', methods=['POST'])
def add_student():
    name = request.form.get('name', '').strip()
    roll_number = request.form.get('rollNumber', '').strip()
    phone_number = request.form.get('phoneNumber', '').strip()
    photo = request.files.get('photo')

    if not name or not roll_number or not photo:
        return jsonify({'error': 'Name, roll number, and photo are required'}), 400

    student_id = uuid.uuid4().hex
    extension = os.path.splitext(secure_filename(photo.filename or 'student.jpg'))[1].lower() or '.jpg'
    photo_filename = f"{student_id}{extension}"
    photo_path = os.path.join(STUDENT_PHOTO_DIR, photo_filename)
    photo.save(photo_path)

    students = load_students()
    student = {
        'id': student_id,
        'name': name,
        'rollNumber': roll_number,
        'phoneNumber': phone_number,
        'photoFilename': photo_filename,
        'createdAt': datetime.now().strftime('%Y-%m-%d'),
    }
    students.append(student)
    save_students(students)

    return jsonify(serialize_student(student))


@app.route('/api/students/<student_id>', methods=['PUT', 'PATCH'])
def update_student(student_id):
    students = load_students()
    student_index = next((index for index, item in enumerate(students) if item.get('id') == student_id), None)

    if student_index is None:
        return jsonify({'error': 'Student not found'}), 404

    existing_student = students[student_index]
    name = request.form.get('name', existing_student.get('name', '')).strip()
    roll_number = request.form.get('rollNumber', existing_student.get('rollNumber', '')).strip()
    phone_number = request.form.get('phoneNumber', existing_student.get('phoneNumber', '')).strip()
    photo = request.files.get('photo')

    if not name or not roll_number:
        return jsonify({'error': 'Name and roll number are required'}), 400

    photo_filename = existing_student.get('photoFilename', '')

    if photo:
        if photo_filename:
            old_photo_path = os.path.join(STUDENT_PHOTO_DIR, photo_filename)
            if os.path.exists(old_photo_path):
                os.remove(old_photo_path)
        extension = os.path.splitext(secure_filename(photo.filename or 'student.jpg'))[1].lower() or '.jpg'
        photo_filename = f"{student_id}{extension}"
        photo_path = os.path.join(STUDENT_PHOTO_DIR, photo_filename)
        photo.save(photo_path)

    updated_student = {
        **existing_student,
        'name': name,
        'rollNumber': roll_number,
        'phoneNumber': phone_number,
        'photoFilename': photo_filename,
    }
    students[student_index] = updated_student
    save_students(students)

    return jsonify(serialize_student(updated_student))


@app.route('/api/students/<student_id>', methods=['DELETE'])
def delete_student(student_id):
    students = load_students()
    remaining = []
    removed = False

    for student in students:
        if student.get('id') == student_id:
            removed = True
            photo_filename = student.get('photoFilename', '')
            if photo_filename:
                photo_path = os.path.join(STUDENT_PHOTO_DIR, photo_filename)
                if os.path.exists(photo_path):
                    os.remove(photo_path)
            continue
        remaining.append(student)

    if not removed:
        return jsonify({'error': 'Student not found'}), 404

    save_students(remaining)
    return jsonify({'success': True})


@app.route('/api/teacher-credentials', methods=['GET'])
def get_teacher_credentials():
    return jsonify(load_teacher_profile())


@app.route('/api/teacher-credentials', methods=['PUT', 'PATCH'])
def update_teacher_credentials():
    data = request.get_json(silent=True) or {}
    if not isinstance(data, dict):
        data = {}

    profile = load_teacher_profile()
    profile.update({
        'email': str(data.get('email', profile.get('email', ''))).strip().lower(),
        'password': str(data.get('password', profile.get('password', ''))).strip(),
        'name': str(data.get('name', profile.get('name', ''))).strip(),
        'className': str(data.get('className', profile.get('className', ''))).strip(),
        'section': str(data.get('section', profile.get('section', ''))).strip(),
    })

    if 'classStartDate' in data:
        profile['classStartDate'] = str(data.get('classStartDate', '') or '').strip()

    if 'dayOverrides' in data:
        overrides = data.get('dayOverrides', {}) or {}
        profile['dayOverrides'] = {str(date_key): str(status_value) for date_key, status_value in overrides.items() if str(date_key).strip()}

    save_teacher_profile(profile)
    return jsonify(profile)


@app.route('/api/student-photo/<filename>')
def get_student_photo(filename):
    return send_file(os.path.join(STUDENT_PHOTO_DIR, filename))


@app.route('/api/attendance/start', methods=['POST'])
def start_attendance():
    global attendance_process, current_excel_file, attendance_last_error

    data = request.get_json(silent=True) or {}
    mode = data.get('mode', 'webcam')
    source = normalize_source(mode, data.get('ipAddress') or data.get('source') or '')

    if attendance_process and attendance_process.poll() is None:
        return jsonify({'error': 'Attendance already running'}), 400

    if not os.path.exists(ATTENDANCE_SCRIPT):
        return jsonify({'error': 'Attendance runner is missing'}), 500

    try:
        students = load_students()
        photo_ready_students, missing_students = collect_photo_ready_students(students)

        if not students:
            return jsonify({'error': 'No students found. Add students before starting attendance.'}), 400

        if not photo_ready_students:
            attendance_last_error = 'No student photos were found. Upload a photo for each student in Students before starting attendance.'
            return jsonify({'error': attendance_last_error}), 400

        current_excel_file = create_session_excel_file()
        clear_latest_frame()
        if mode == 'webcam':
            source = '0'
        elif not source:
            return jsonify({'error': 'IP camera or CCTV source is required'}), 400

        env = os.environ.copy()
        env['ATTENDANCE_EXPORT_FILE'] = current_excel_file
        env['ATTENDANCE_REPORT_TITLE'] = 'Attenzo Attendance Report'
        env['ATTENDANCE_SHOW_WINDOW'] = '0'
        env['ATTENDANCE_USE_LOCAL_FILES'] = '1'
        env['ATTENDANCE_RUNTIME_DIR'] = ATTENDANCE_RUNTIME_DIR
        env['ATTENDANCE_FRAME_FILE'] = ATTENDANCE_FRAME_FILE
        env['ATTENDANCE_SESSION_ROLLS'] = ','.join(
            [str(student.get('rollNumber', '')).strip() for student in students if student.get('rollNumber')]
        )
        env['ATTENDANCE_STUDENTS_API_URL'] = f'{request.host_url.rstrip("/")}/api/students'

        attendance_process = subprocess.Popen(
            [sys.executable, '-u', ATTENDANCE_SCRIPT, '--mode', mode, '--source', source],
            cwd=BASE_DIR,
            env=env,
            preexec_fn=os.setsid,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )

        attendance_last_error = ''
        monitor_thread = threading.Thread(target=monitor_attendance_process, args=(attendance_process,), daemon=True)
        monitor_thread.start()

        return jsonify({
            'success': True,
            'message': f'Attendance started in {mode} mode',
            'mode': mode,
            'source': source,
            'excelFile': current_excel_file,
        })
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


@app.route('/api/attendance/stop', methods=['POST'])
def stop_attendance():
    global attendance_process, attendance_last_error

    if attendance_process and attendance_process.poll() is None:
        try:
            os.killpg(os.getpgid(attendance_process.pid), signal.SIGTERM)
        except Exception:
            attendance_process.terminate()
        try:
            attendance_process.wait(timeout=10)
        except Exception:
            try:
                os.killpg(os.getpgid(attendance_process.pid), signal.SIGKILL)
            except Exception:
                attendance_process.kill()
        attendance_process = None
        attendance_last_error = ''
        clear_latest_frame()
        return jsonify({
            'success': True,
            'message': 'Attendance stopped',
            'excelUrl': '/api/reports/excel',
        })

    attendance_process = None
    attendance_last_error = ''
    clear_latest_frame()
    return jsonify({'error': 'No attendance process running'}), 400


@app.route('/api/attendance/status')
def attendance_status():
    running = attendance_process is not None and attendance_process.poll() is None
    present = []

    excel_file = get_active_excel_file()
    sheet_name = get_today_sheet_name()

    if os.path.exists(excel_file):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(excel_file)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=3, values_only=True):
                    record = parse_attendance_row(row)
                    if record and record.get('status') == 'Present':
                        present.append(record.get('name'))
        except Exception:
            pass

    absent = []
    present_names = set(present)
    for student in load_students():
        if student.get('name') not in present_names:
            absent.append(student.get('name'))

    return jsonify({
        'running': running,
        'present': present,
        'absent': absent,
        'message': '' if running else attendance_last_error,
    })


@app.route('/api/attendance/feed')
def attendance_feed():
    running = attendance_process is not None and attendance_process.poll() is None
    if not running:
        return jsonify({'error': 'No attendance process running'}), 400

    return Response(
        stream_latest_frame(),
        mimetype='multipart/x-mixed-replace; boundary=frame',
        headers={
            'Cache-Control': 'no-cache, no-store, must-revalidate',
            'Pragma': 'no-cache',
            'Expires': '0',
        },
    )


@app.route('/api/attendance/settings', methods=['GET'])
def get_attendance_settings():
    profile = load_teacher_profile()
    return jsonify({
        'classStartDate': profile.get('classStartDate', ''),
        'dayOverrides': get_day_overrides(profile),
    })


@app.route('/api/attendance/settings', methods=['PUT', 'PATCH'])
def update_attendance_settings():
    data = request.get_json(silent=True) or {}
    if not isinstance(data, dict):
        data = {}

    profile = load_teacher_profile()
    if 'classStartDate' in data:
        profile['classStartDate'] = str(data.get('classStartDate', '') or '').strip()

    if 'dayOverrides' in data:
        overrides = data.get('dayOverrides', {}) or {}
        profile['dayOverrides'] = {str(date_key): str(status_value) for date_key, status_value in overrides.items() if str(date_key).strip()}

    save_teacher_profile(profile)
    return jsonify({
        'classStartDate': profile.get('classStartDate', ''),
        'dayOverrides': get_day_overrides(profile),
    })


@app.route('/api/attendance/mark-day', methods=['POST'])
def mark_attendance_day():
    data = request.get_json(silent=True) or {}
    if not isinstance(data, dict):
        data = {}

    day_date = str(data.get('date', '') or '').strip()
    status = str(data.get('status', 'Holiday') or '').strip()
    if not day_date:
        return jsonify({'error': 'A date is required'}), 400

    if status not in {'Holiday', 'Sunday'}:
        return jsonify({'error': 'Status must be Holiday or Sunday'}), 400

    try:
        datetime.strptime(day_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'Use a YYYY-MM-DD date format'}), 400

    students = load_students()
    if not students:
        return jsonify({'error': 'No students found'}), 400

    excel_file = get_active_excel_file() or create_session_excel_file()
    if not persist_day_override(excel_file, day_date, status, students):
        return jsonify({'error': 'Unable to save the day override'}), 400

    profile = load_teacher_profile()
    overrides = get_day_overrides(profile)
    overrides[day_date] = status
    profile['dayOverrides'] = overrides
    save_teacher_profile(profile)

    return jsonify({
        'success': True,
        'message': f'{status} marked for {day_date}',
        'date': day_date,
        'status': status,
    })


@app.route('/api/reports')
def get_reports():
    return jsonify(collect_attendance_records())


@app.route('/api/reports/excel')
def download_excel():
    excel_file = get_active_excel_file()
    if os.path.exists(excel_file):
        return send_file(excel_file, as_attachment=True, download_name=os.path.basename(excel_file))
    return jsonify({'error': 'No attendance data found'}), 404


@app.route('/api/dashboard/stats')
def get_dashboard_stats():
    total_students = len(load_students())
    today_attendance = 0
    all_records = collect_attendance_records()
    today_key = get_today_sheet_name()

    for record in all_records:
        if record.get('date') == today_key and normalize_status_value(record.get('status')) == 'present':
            today_attendance += 1

    return jsonify({
        'totalStudents': total_students,
        'activeCameras': 1 if attendance_process and attendance_process.poll() is None else 0,
        'todayAttendance': today_attendance,
        'totalAlerts': 0,
    })


if __name__ == '__main__':
    print('CCTV Attendance API Server starting...')
    print(f'Student photos: {STUDENT_PHOTO_DIR}')
    print(f'Students file: {STUDENTS_FILE}')
    print(f'Excel file: {EXCEL_FILE}')
    app.run(host=HOST, port=PORT, debug=False, use_reloader=False)
