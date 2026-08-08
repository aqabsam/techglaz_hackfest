import argparse
import json
import os
import signal
import time
import urllib.request
from datetime import datetime

import cv2
import face_recognition
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STUDENTS_FILE = os.path.join(BASE_DIR, 'students.json')
EXPORT_DIR = os.environ.get('ATTENDANCE_EXPORT_DIR', os.path.join(BASE_DIR, 'exports'))
os.makedirs(EXPORT_DIR, exist_ok=True)
DEFAULT_EXCEL_FILE = os.environ.get('ATTENDANCE_EXPORT_FILE') or os.path.join(EXPORT_DIR, 'Attendance.xlsx')
EXCEL_FILE = os.environ.get('ATTENDANCE_EXPORT_FILE', DEFAULT_EXCEL_FILE)
REPORT_TITLE = os.environ.get('ATTENDANCE_REPORT_TITLE', 'Attenzo Attendance Report')
STUDENTS_API_URL = os.environ.get('ATTENDANCE_STUDENTS_API_URL', 'http://localhost:5000/api/students')
API_BASE_URL = os.environ.get('ATTENDANCE_API_BASE_URL') or STUDENTS_API_URL.rsplit('/api/students', 1)[0]
STOP_REQUESTED = False
ACTIVE_STUDENTS = []
WEBCAM_OPEN_TIMEOUT_SECONDS = float(os.environ.get('ATTENDANCE_CAMERA_OPEN_TIMEOUT_SECONDS', '12'))


hidden_name_fragments = ['mohd aqab sami', 'mohd aqabsami', 'aban shami']
hidden_phone_digits = {'8986392298'}


def normalize_student_value(value):
    return str(value or '').strip().lower().replace('\n', ' ').replace('\t', ' ')


def digits_only(value):
    return ''.join(ch for ch in str(value or '') if ch.isdigit())


def is_visible_student(student):
    normalized_name = normalize_student_value(student.get('name', ''))
    normalized_phone = digits_only(student.get('phoneNumber', ''))
    return not any(fragment in normalized_name for fragment in hidden_name_fragments) and normalized_phone not in hidden_phone_digits


def filter_visible_students(students):
    return [student for student in students if is_visible_student(student)]


def roll_sort_key(student):
    roll_number = str(student.get('rollNumber', '')).strip()
    digits = ''.join(ch for ch in roll_number if ch.isdigit())
    if digits:
        return (0, int(digits), roll_number.lower())
    return (1, roll_number.lower())


def sorted_students_by_roll(students):
    return sorted(students, key=roll_sort_key)


def student_attendance_key(student):
    roll_number = str(student.get('rollNumber', '')).strip()
    student_id = str(student.get('id', '')).strip()
    return roll_number or student_id


def today_sheet_name():
    return datetime.now().strftime('%Y-%m-%d')


def today_sheet_title():
    now = datetime.now()
    return f"Attendance - {now.strftime('%Y-%m-%d')} ({now.strftime('%A')})"


def build_sheet_headers(ws):
    ws.merge_cells("A1:D1")
    ws["A1"] = today_sheet_title()
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="0F766E")

    ws.append(["Roll No", "Name", "Check In Time", "Status"])

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col in ["A", "B", "C", "D"]:
        ws[f"{col}2"].font = Font(bold=True, color="FFFFFF")
        ws[f"{col}2"].alignment = Alignment(horizontal="center")
        ws[f"{col}2"].border = thin_border
        ws[f"{col}2"].fill = PatternFill("solid", fgColor="0891B2")
        ws.column_dimensions[col].width = 24


def ensure_today_sheet(workbook):
    sheet_name = today_sheet_name()
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
    else:
        if workbook.sheetnames == ['Sheet'] and workbook['Sheet'].max_row == 1 and workbook['Sheet']['A1'].value is None:
            workbook.remove(workbook['Sheet'])
        ws = workbook.create_sheet(title=sheet_name)
        build_sheet_headers(ws)
        return ws

    if ws["A1"].value is None:
        build_sheet_headers(ws)
    return ws


def sheet_has_records(ws):
    return ws.max_row >= 3 and any(ws[f"A{row}"].value for row in range(3, ws.max_row + 1))


def load_today_marks(ws):
    marks = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue

        roll_number = str(row[0] or '').strip()
        name = str(row[1] or '').strip()
        checkin_time = str(row[2] or '').strip()
        status = str(row[3] or '').strip()

        if roll_number:
            marks[roll_number] = {
                'name': name,
                'status': status,
                'checkinTime': checkin_time,
            }

    return marks


def load_existing_today_attendance():
    if not os.path.exists(EXCEL_FILE):
        return {}

    try:
        wb = load_workbook(EXCEL_FILE)
        sheet_name = today_sheet_name()
        if sheet_name not in wb.sheetnames:
            return {}

        ws = wb[sheet_name]
        existing = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 4:
                continue

            roll_number = str(row[0] or '').strip()
            name = str(row[1] or '').strip()
            checkin_time = str(row[2] or '').strip()
            status = str(row[3] or '').strip()

            if roll_number and status == 'Present':
                existing[roll_number] = {
                    'name': name,
                    'rollNumber': roll_number,
                    'status': status,
                    'time': checkin_time,
                }

        return existing
    except Exception:
        return {}


def rebuild_today_sheet(ws, known_students, attendance):
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)

    absent_students = []

    for student in sorted_students_by_roll(known_students):
        roll_number = str(student.get('rollNumber', '')).strip()
        if not roll_number:
            continue

        attendance_record = attendance.get(roll_number)
        if attendance_record:
            status = 'Present'
            checkin_time = attendance_record.get('time', 'N/A') or 'N/A'
        else:
            status = 'Absent'
            checkin_time = 'N/A'
            absent_students.append(student)

        ws.append([roll_number, student.get('name', ''), checkin_time, status])

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for row_num in range(3, ws.max_row + 1):
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row_num}"].alignment = Alignment(horizontal="center")
            ws[f"{col}{row_num}"].border = thin_border
            if row_num % 2 == 0:
                ws[f"{col}{row_num}"].fill = PatternFill("solid", fgColor="E0F2FE")
            if ws[f"C{row_num}"].value == 'Absent':
                ws[f"C{row_num}"].fill = PatternFill("solid", fgColor="FEE2E2")
            elif ws[f"C{row_num}"].value == 'Present':
                ws[f"C{row_num}"].fill = PatternFill("solid", fgColor="DCFCE7")

    return absent_students


def build_absent_sheet(workbook, absent_students):
    sheet_name = f"Absent-{today_sheet_name()}"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    ws = workbook.create_sheet(title=sheet_name)
    ws.merge_cells("A1:B1")
    ws["A1"] = f"Absent Students - {datetime.now().strftime('%Y-%m-%d')} ({datetime.now().strftime('%A')})"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="7F1D1D")

    ws.append(["Name", "Roll Number"])

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col in ["A", "B"]:
        ws[f"{col}2"].font = Font(bold=True, color="FFFFFF")
        ws[f"{col}2"].alignment = Alignment(horizontal="center")
        ws[f"{col}2"].border = thin_border
        ws[f"{col}2"].fill = PatternFill("solid", fgColor="B91C1C")
        ws.column_dimensions[col].width = 28

    for student in sorted_students_by_roll(absent_students):
        ws.append([
            student.get('name', ''),
            str(student.get('rollNumber', '')).strip(),
        ])

    for row_num in range(3, ws.max_row + 1):
        for col in ["A", "B"]:
            ws[f"{col}{row_num}"].alignment = Alignment(horizontal="center")
            ws[f"{col}{row_num}"].border = thin_border
            if row_num % 2 == 0:
                ws[f"{col}{row_num}"].fill = PatternFill("solid", fgColor="FEE2E2")


def sync_today_workbook(known_students, attendance):
    if not known_students:
        return

    ensure_workbook()

    wb = load_workbook(EXCEL_FILE)
    ws = ensure_today_sheet(wb)
    absent_students = rebuild_today_sheet(ws, known_students, attendance)
    build_absent_sheet(wb, absent_students)
    wb.save(EXCEL_FILE)


def handle_stop_signal(signum, frame):
    global STOP_REQUESTED
    STOP_REQUESTED = True


signal.signal(signal.SIGTERM, handle_stop_signal)
signal.signal(signal.SIGINT, handle_stop_signal)


def ensure_workbook():
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()

    ensure_today_sheet(wb)
    wb.save(EXCEL_FILE)


def load_students():
    try:
        with urllib.request.urlopen(STUDENTS_API_URL, timeout=10) as response:
            body = response.read().decode('utf-8').strip()
            if not body:
                raise ValueError('Empty student payload')
            data = json.loads(body)
            students = filter_visible_students(data if isinstance(data, list) else [])
            if students:
                return students
            raise ValueError('No visible students returned from API')
    except Exception:
        if not os.path.exists(STUDENTS_FILE):
            return []

        try:
            with open(STUDENTS_FILE, 'r', encoding='utf-8') as file:
                data = json.load(file)
                return filter_visible_students(data if isinstance(data, list) else [])
        except Exception:
            return []


def fetch_image_bytes(photo_url):
    with urllib.request.urlopen(photo_url, timeout=10) as response:
        return response.read()


def student_photo_source(student):
    photo_url = student.get('photoUrl') or ''
    if photo_url:
        return photo_url

    photo_filename = student.get('photoFilename') or ''
    if photo_filename:
        return f'{API_BASE_URL}/api/student-photo/{photo_filename}'

    return ''


def build_face_database(students):
    known_encodings = []
    known_students = []

    for student in students:
        photo_source = student_photo_source(student)
        if not photo_source:
            continue

        try:
            image_bytes = fetch_image_bytes(photo_source)
        except Exception:
            continue

        image_array = np.asarray(bytearray(image_bytes), dtype=np.uint8)
        image = cv2.imdecode(image_array, cv2.IMREAD_COLOR)
        if image is None:
            continue

        rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        encodings = face_recognition.face_encodings(rgb_image)
        if not encodings:
            continue

        known_encodings.append(encodings[0])
        known_students.append(student)

    return known_encodings, known_students


def mark_attendance(student, attendance):
    name = student.get('name', 'Unknown')
    roll_number = str(student.get('rollNumber', '')).strip()
    attendance_key = student_attendance_key(student)
    if not attendance_key or attendance_key in attendance:
        return

    ensure_workbook()

    now = datetime.now()
    time_string = now.strftime('%H:%M:%S')

    attendance[attendance_key] = {
        "rollNumber": roll_number,
        "time": time_string,
        "status": "Present",
        "name": name,
    }

    sync_today_workbook(ACTIVE_STUDENTS, attendance)
    print(f"Checked in: {name} at {time_string}")


def resolve_source(mode, source_text):
    if mode == 'webcam':
        return 0

    if source_text.isdigit():
        return int(source_text)

    return source_text


def open_camera_capture(source):
    candidate_sources = [source]

    if isinstance(source, int):
        candidate_sources = [source, 0, 1, 2, 3]

    deadline = time.time() + WEBCAM_OPEN_TIMEOUT_SECONDS if isinstance(source, int) else time.time()

    while True:
        for candidate in candidate_sources:
            if isinstance(candidate, int):
                for backend in (cv2.CAP_AVFOUNDATION, cv2.CAP_ANY):
                    capture = cv2.VideoCapture(candidate, backend)
                    if capture.isOpened():
                        return capture
                    capture.release()
            else:
                capture = cv2.VideoCapture(candidate, cv2.CAP_ANY)
                if capture.isOpened():
                    return capture
                capture.release()

        if not isinstance(source, int) or time.time() >= deadline:
            break

        time.sleep(0.5)

    return cv2.VideoCapture(source)


def main():
    parser = argparse.ArgumentParser(description='CCTV Attendance Runner')
    parser.add_argument('--mode', default='webcam', choices=['webcam', 'ip_camera', 'cctv'])
    parser.add_argument('--source', default='0')
    args = parser.parse_args()

    ensure_workbook()
    global ACTIVE_STUDENTS
    students = load_students()
    ACTIVE_STUDENTS = students

    if not students:
        print("No students found. Add student records before starting attendance.")
        return

    known_encodings, known_students = build_face_database(students)
    if not known_encodings:
        print("No valid face encodings found. Check student photos.")
        return

    source = resolve_source(args.mode, args.source)
    video = open_camera_capture(source)

    if not video.isOpened():
        print(f"Unable to open video source: {source}")
        return

    attendance = load_existing_today_attendance()
    last_seen = {}

    print(f"Starting attendance in {args.mode} mode from {source}")

    try:
        while True:
            if STOP_REQUESTED:
                print("Stop requested.")
                break

            ret, frame = video.read()
            if not ret:
                print("Video stream ended or could not be read.")
                break

            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

            face_locations = face_recognition.face_locations(rgb_small)
            face_encodings = face_recognition.face_encodings(rgb_small, face_locations)

            for face_encoding, face_location in zip(face_encodings, face_locations):
                matches = face_recognition.compare_faces(known_encodings, face_encoding)
                face_distances = face_recognition.face_distance(known_encodings, face_encoding)

                if len(face_distances) == 0:
                    continue

                best_match_index = int(np.argmin(face_distances))
                student = known_students[best_match_index]

                if matches[best_match_index] and face_distances[best_match_index] < 0.5:
                    mark_attendance(student, attendance)
                    name = student.get('name', 'Unknown')
                    last_seen[name] = time.time()
                    attendance_key = student_attendance_key(student)

                    top, right, bottom, left = face_location
                    top *= 4
                    right *= 4
                    bottom *= 4
                    left *= 4
                    cv2.rectangle(frame, (left, top), (right, bottom), (0, 220, 120), 2)
                    cv2.putText(frame, f"Roll {student.get('rollNumber', '')}", (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 220, 120), 2)

                    if attendance_key in attendance:
                        cv2.putText(
                            frame,
                            f"Checked in at {attendance[attendance_key]['time']}",
                            (left, bottom + 24),
                            cv2.FONT_HERSHEY_SIMPLEX,
                            0.6,
                            (255, 255, 255),
                            2,
                        )
                else:
                    top, right, bottom, left = face_location
                    top *= 4
                    right *= 4
                    bottom *= 4
                    left *= 4
                    cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
                    cv2.putText(frame, "Unknown", (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)

            present_students = [
                student.get('name', '')
                for student in known_students
                if student_attendance_key(student) in attendance
            ]
            absent_students = [
                student.get('name', '')
                for student in known_students
                if student_attendance_key(student) not in attendance
            ]

            cv2.rectangle(frame, (0, 0), (frame.shape[1], 64), (15, 23, 42), -1)
            cv2.putText(frame, "CCTV Attendance Running", (18, 28), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
            cv2.putText(frame, f"Present: {len(present_students)}", (18, 52), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (34, 197, 94), 2)
            cv2.putText(frame, f"Absent: {len(absent_students)}", (180, 52), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (248, 113, 113), 2)

            y_offset = 100
            cv2.putText(frame, "Present", (20, y_offset), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (34, 197, 94), 2)
            for index, student_name in enumerate(present_students[:8], start=1):
                cv2.putText(frame, student_name, (20, y_offset + index * 26), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (34, 197, 94), 2)

            x_offset = frame.shape[1] - 260
            cv2.putText(frame, "Absent", (x_offset, y_offset), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (248, 113, 113), 2)
            for index, student_name in enumerate(absent_students[:8], start=1):
                cv2.putText(frame, student_name, (x_offset, y_offset + index * 26), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (248, 113, 113), 2)

            cv2.imshow("CCTV Attendance", frame)

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
    finally:
        if os.path.exists(EXCEL_FILE):
            try:
                sync_today_workbook(known_students, attendance)
            except Exception:
                pass
        video.release()
        cv2.destroyAllWindows()
        print(f"Attendance saved to {EXCEL_FILE}")


if __name__ == '__main__':
    main()
