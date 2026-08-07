# cd ~/Desktop
# python3 -m venv venv
# source venv/bin/activate
# python3 Attendances.py
# cd ~/Desktop
# python3 -m venv venv
# source venv/bin/activate
# python3 Attendance.py

import cv2
import face_recognition
import numpy as np
import os
from datetime import datetime
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# =======================
# 1️⃣ Path to save attendance Excel file on Desktop
# =======================
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
excel_file = os.path.join(desktop_path, "Attendance.xlsx")

# =======================
# 2️⃣ Create workbook and headers if not exists
# =======================
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Title row
    ws.merge_cells("A1:E1")
    ws["A1"] = "ADVANCED BIOMETRIC IN CLASSES USING CCTV CAMERA"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="4F81BD")  # Blue background

    # Headers
    headers = ["Name", "Check-in Time", "Status", "Date", "Day"]
    ws.append(headers)

    # Style headers
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for col, header in zip(["A","B","C","D","E"], headers):
        ws[f"{col}2"].font = Font(bold=True, color="FFFFFF")
        ws[f"{col}2"].alignment = Alignment(horizontal="center")
        ws[f"{col}2"].border = thin_border
        ws[f"{col}2"].fill = PatternFill("solid", fgColor="4BACC6")  # Teal header
        ws.column_dimensions[col].width = 22

    wb.save(excel_file)

# =======================
# 3️⃣ Load known faces
# =======================
path = "KnownFaces"
images = []
student_names = []

if not os.path.exists(path):
    print(f"❌ Folder '{path}' not found! Please create it and add student images.")
    exit()

for filename in os.listdir(path):
    img = cv2.imread(os.path.join(path, filename))
    if img is not None:
        images.append(img)
        student_names.append(os.path.splitext(filename)[0])  # Example: "John"

# =======================
# 4️⃣ Encode known faces
# =======================
known_encodings = []
for img in images:
    rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    enc = face_recognition.face_encodings(rgb_img)[0]
    known_encodings.append(enc)

# =======================
# 5️⃣ Attendance tracking
# =======================
attendance = {}
checkin_display = {}

def mark_attendance(name):
    """Mark attendance for recognized student and save to Excel"""
    if name not in attendance:
        now = datetime.now()
        time_string = now.strftime('%H:%M:%S')
        date_string = now.strftime('%Y-%m-%d')
        day_string = now.strftime('%A')

        attendance[name] = {"time": time_string, "status": "Present"}
        checkin_display[name] = time.time()
        print(f"✅ {name} checked in at {time_string} on {date_string} ({day_string})")

        # Append to Excel
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append([name, time_string, "Present", date_string, day_string])

        # Apply border and alignment
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        row_num = ws.max_row
        for col in ["A","B","C","D","E"]:
            ws[f"{col}{row_num}"].alignment = Alignment(horizontal="center")
            ws[f"{col}{row_num}"].border = thin_border

        # Alternate row color for readability
        if row_num % 2 == 1:
            for col in ["A","B","C","D","E"]:
                ws[f"{col}{row_num}"].fill = PatternFill("solid", fgColor="D9E1F2")  # light blue

        wb.save(excel_file)

# =======================
# 6️⃣ Video Source Setup
# =======================
# Options:
#  - Local webcam -> 0
#  - Mobile camera via IP Webcam app -> "http://<phone-ip>:8080/video"
#  - CCTV Camera RTSP -> "rtsp://user:pass@IP:554/stream"

# Uncomment the one you want:
# VIDEO_SOURCE = 0  # laptop webcam
VIDEO_SOURCE = "http://192.168.147.103:4747/video"  # mobile IP webcam
# VIDEO_SOURCE = "rtsp://user:pass@192.168.1.50:554/stream1"  # example CCTV

video = cv2.VideoCapture(VIDEO_SOURCE)

if not video.isOpened():
    print("⚠️ Unable to open video source. Check VIDEO_SOURCE!")
    exit()

# =======================
# 7️⃣ Main Loop
# =======================
while True:
    ret, frame = video.read()
    if not ret:
        print("⚠️ Unable to read video stream. Check VIDEO_SOURCE!")
        break

    small_frame = cv2.resize(frame, (0,0), fx=0.25, fy=0.25)
    rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    face_locations = face_recognition.face_locations(rgb_small)
    face_encodings = face_recognition.face_encodings(rgb_small, face_locations)

    for face_encoding, face_location in zip(face_encodings, face_locations):
        matches = face_recognition.compare_faces(known_encodings, face_encoding)
        face_distances = face_recognition.face_distance(known_encodings, face_encoding)
        best_match_index = np.argmin(face_distances)

        if matches[best_match_index] and face_distances[best_match_index] < 0.5:
            name = student_names[best_match_index]
            mark_attendance(name)

            # Draw green rectangle + label
            top, right, bottom, left = face_location
            top*=4; right*=4; bottom*=4; left*=4
            cv2.rectangle(frame, (left, top), (right, bottom), (0,255,0), 2)
            cv2.putText(frame, name, (left, top-10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0,255,0), 2)

            if name in checkin_display and time.time()-checkin_display[name]<3:
                cv2.putText(frame, f"Checked in at {attendance[name]['time']}",
                            (left, bottom+30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255,255,0),2)
        else:
            # Unknown face
            top,right,bottom,left = face_location
            top*=4; right*=4; bottom*=4; left*=4
            cv2.rectangle(frame, (left,top),(right,bottom),(0,0,255),2)
            cv2.putText(frame, "Unknown",(left,top-10),cv2.FONT_HERSHEY_SIMPLEX,0.9,(0,0,255),2)

    # Show absent students (left)
    absent_students = [s for s in student_names if s not in attendance]
    y_offset = 50
    cv2.putText(frame,"Absent Students:",(10,y_offset),cv2.FONT_HERSHEY_SIMPLEX,0.9,(0,0,255),2)
    for i,student in enumerate(absent_students,start=1):
        cv2.putText(frame,student,(10,y_offset+i*30),cv2.FONT_HERSHEY_SIMPLEX,0.8,(0,0,255),2)

    # Show present students (right)
    present_students = list(attendance.keys())
    if present_students:
        y_offset = 50
        x_offset = frame.shape[1]-300
        cv2.putText(frame,"Present Students:",(x_offset,y_offset),cv2.FONT_HERSHEY_SIMPLEX,0.9,(0,200,0),2)
        for i,student in enumerate(present_students,start=1):
            cv2.putText(frame,student,(x_offset,y_offset+i*30),cv2.FONT_HERSHEY_SIMPLEX,0.8,(0,200,0),2)

    cv2.imshow("CCTV Biometric Attendance", frame)
    if cv2.waitKey(1)&0xFF==ord('q'):
        break

video.release()
cv2.destroyAllWindows()
print(f"📌 Attendance saved to {excel_file}")
